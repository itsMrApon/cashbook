import os
import uuid
from datetime import datetime
from decimal import Decimal
from flask import current_app
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json

def allowed_file(filename):
    """Check if file extension is allowed"""
    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file):
    """Save uploaded file and return filename"""
    if file and allowed_file(file.filename):
        # Generate unique filename
        filename = str(uuid.uuid4()) + '.' + file.filename.rsplit('.', 1)[1].lower()
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return filename
    return None

def delete_file(filename):
    """Delete file from upload folder"""
    try:
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
    except Exception as e:
        current_app.logger.error(f"Error deleting file {filename}: {str(e)}")
    return False

def format_currency(amount):
    """Format amount as currency"""
    if amount is None:
        return "$0.00"
    return f"${amount:,.2f}"

def parse_tags(tag_string):
    """Parse comma-separated tag string into list"""
    if not tag_string:
        return []
    return [tag.strip() for tag in tag_string.split(',') if tag.strip()]

def generate_pdf_report(transactions, title="Transaction Report", user_filter=None, date_range=None):
    """Generate PDF report from transactions"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph(title, title_style))
    
    # Report info
    info_style = styles['Normal']
    if user_filter:
        elements.append(Paragraph(f"User: {user_filter}", info_style))
    if date_range:
        elements.append(Paragraph(f"Period: {date_range}", info_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", info_style))
    elements.append(Spacer(1, 20))
    
    # Summary
    total_income = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')
    net_amount = total_income - total_expense
    
    summary_data = [
        ['Summary', ''],
        ['Total Income', format_currency(total_income)],
        ['Total Expenses', format_currency(total_expense)],
        ['Net Amount', format_currency(net_amount)]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Transactions table
    if transactions:
        data = [['Date', 'Type', 'Description', 'Category', 'Amount']]
        
        for transaction in transactions:
            data.append([
                transaction.transaction_date.strftime('%Y-%m-%d'),
                transaction.type.title(),
                transaction.description[:50] + ('...' if len(transaction.description) > 50 else ''),
                transaction.category.name,
                format_currency(transaction.amount)
            ])
        
        table = Table(data, colWidths=[1.2*inch, 1*inch, 2.5*inch, 1.5*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),  # Amount column right-aligned
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(Paragraph("Transaction Details", styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(transactions, title="Transaction Report"):
    """Generate Excel report from transactions"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ['Date', 'Type', 'Description', 'Category', 'Party', 'Amount', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, transaction in enumerate(transactions, 4):
        ws.cell(row=row, column=1, value=transaction.transaction_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=transaction.type.title())
        ws.cell(row=row, column=3, value=transaction.description)
        ws.cell(row=row, column=4, value=transaction.category.name)
        ws.cell(row=row, column=5, value=transaction.party or '')
        ws.cell(row=row, column=6, value=float(transaction.amount))
        ws.cell(row=row, column=7, value=transaction.notes or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "Financial Summary"
    summary_ws['A1'].font = Font(size=14, bold=True)
    
    total_income = sum(t.amount for t in transactions if t.type == 'income')
    total_expense = sum(t.amount for t in transactions if t.type == 'expense')
    net_amount = total_income - total_expense
    
    summary_data = [
        ['Total Income', float(total_income)],
        ['Total Expenses', float(total_expense)],
        ['Net Amount', float(net_amount)]
    ]
    
    for row, (label, value) in enumerate(summary_data, 3):
        summary_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=row, column=2, value=value)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def log_audit_action(user_id, action, table_name=None, record_id=None, old_values=None, new_values=None):
    """Log audit action"""
    from app.models import AuditLog
    from app import db
    from flask import request
    
    try:
        audit_log = AuditLog(
            user_id=user_id,
            action=action,
            table_name=table_name,
            record_id=record_id,
            old_values=json.dumps(old_values) if old_values else None,
            new_values=json.dumps(new_values) if new_values else None,
            ip_address=request.remote_addr if request else None,
            user_agent=request.headers.get('User-Agent') if request else None
        )
        db.session.add(audit_log)
        db.session.commit()
    except Exception as e:
        current_app.logger.error(f"Error logging audit action: {str(e)}")

def get_dashboard_stats(user=None):
    """Get dashboard statistics"""
    from app.models import Transaction, User
    from sqlalchemy import func
    from datetime import datetime, timedelta
    
    # Base query
    query = Transaction.query
    if user and not user.has_permission('manage_users'):
        query = query.filter(Transaction.user_id == user.id)
    
    # This month
    this_month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_income = query.filter(
        Transaction.type == 'income',
        Transaction.transaction_date >= this_month_start.date()
    ).with_entities(func.sum(Transaction.amount)).scalar() or 0
    
    this_month_expenses = query.filter(
        Transaction.type == 'expense',
        Transaction.transaction_date >= this_month_start.date()
    ).with_entities(func.sum(Transaction.amount)).scalar() or 0
    
    # Total counts
    total_transactions = query.count()
    total_users = User.query.filter_by(is_active=True).count()
    
    # Recent transactions
    recent_transactions = query.order_by(Transaction.created_at.desc()).limit(5).all()
    
    return {
        'this_month_income': this_month_income,
        'this_month_expenses': this_month_expenses,
        'this_month_net': this_month_income - this_month_expenses,
        'total_transactions': total_transactions,
        'total_users': total_users,
        'recent_transactions': recent_transactions
    }

def backup_database():
    """Create database backup"""
    # This is a simplified backup - in production, you'd use proper database backup tools
    from app.models import User, Transaction, Category, Tag, SpendingLimit
    from app import db
    
    backup_data = {
        'timestamp': datetime.now().isoformat(),
        'users': [],
        'transactions': [],
        'categories': [],
        'tags': [],
        'spending_limits': []
    }
    
    # Export users (excluding passwords)
    for user in User.query.all():
        backup_data['users'].append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'phone': user.phone,
            'is_active': user.is_active,
            'role_id': user.role_id,
            'created_at': user.created_at.isoformat()
        })
    
    # Export transactions
    for transaction in Transaction.query.all():
        backup_data['transactions'].append({
            'id': transaction.id,
            'type': transaction.type,
            'amount': str(transaction.amount),
            'description': transaction.description,
            'notes': transaction.notes,
            'party': transaction.party,
            'transaction_date': transaction.transaction_date.isoformat(),
            'user_id': transaction.user_id,
            'category_id': transaction.category_id,
            'tags': [tag.name for tag in transaction.tags]
        })
    
    # Export categories
    for category in Category.query.all():
        backup_data['categories'].append({
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'color': category.color,
            'is_system': category.is_system
        })
    
    # Export tags
    for tag in Tag.query.all():
        backup_data['tags'].append({
            'id': tag.id,
            'name': tag.name
        })
    
    # Export spending limits
    for limit in SpendingLimit.query.all():
        backup_data['spending_limits'].append({
            'id': limit.id,
            'type': limit.type,
            'amount': str(limit.amount),
            'is_active': limit.is_active,
            'user_id': limit.user_id,
            'category_id': limit.category_id
        })
    
    return json.dumps(backup_data, indent=2)
